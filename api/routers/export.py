"""Excel export endpoint — generates formatted .xlsx for analyst handoff.

Sheets:
  1. Rankings       — top stations by delay, all 3 lines × morning peak
  2. Anomalies      — flagged stations with severity, actual vs expected
  3. Line Trends    — 30-day daily avg delay per line
  4. Heatmap        — 7×24 weekday×hour delay grid for Dadar CR
  5. Summary        — KPI snapshot (worst station, best line, peak window)

Used by: React frontend export button + VBA macro (excel/RefreshData.bas).
"""

from __future__ import annotations

import datetime
import io
import logging
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from analysis.rankings import peak_rankings
from api.deps import get_store
from pipeline.store import DelayStore

logger = logging.getLogger(__name__)

router = APIRouter(tags=["export"])

# ---------------------------------------------------------------------------
# Colour constants (JPMC-style: dark header, traffic-light data)
# ---------------------------------------------------------------------------
_HDR_FILL = "1C3557"      # dark navy header
_HDR_FONT = "FFFFFF"      # white header text
_RED_FILL = "F4CCCC"      # high severity / high delay
_YELLOW_FILL = "FFF2CC"   # medium severity
_GREEN_FILL = "D9EAD3"    # low delay
_ACCENT = "E63946"        # accent colour (matches React UI)


def _build_workbook(store: DelayStore) -> io.BytesIO:
    """Build the Excel workbook in memory and return a BytesIO buffer."""
    try:
        import openpyxl  # type: ignore[import]
        from openpyxl.formatting.rule import ColorScaleRule  # type: ignore[import]
        from openpyxl.styles import (  # type: ignore[import]
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter  # type: ignore[import]
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed — run: uv add openpyxl") from exc

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── helpers ─────────────────────────────────────────────────────────────

    def hdr_font() -> Font:
        return Font(bold=True, color=_HDR_FONT, size=11)

    def hdr_fill() -> PatternFill:
        return PatternFill("solid", fgColor=_HDR_FILL)

    def thin_border() -> Border:
        s = Side(border_style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def write_header(ws: Any, cols: list[str]) -> None:
        for c, label in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = hdr_font()
            cell.fill = hdr_fill()
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

    def auto_width(ws: Any, min_w: int = 10) -> None:
        for col in ws.columns:
            width = min_w
            for cell in col:
                if cell.value:
                    width = max(width, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    def add_meta(ws: Any) -> None:
        """Stamp generated-at timestamp in bottom-left."""
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1, value=f"Generated: {datetime.datetime.now():%Y-%m-%d %H:%M}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, color="888888", size=9)

    # ── Sheet 1: Rankings ───────────────────────────────────────────────────
    ws_rank = wb.create_sheet("Rankings")
    cols_rank = ["Station", "Line", "Period", "Avg Delay (min)", "CI Lower", "CI Upper"]
    write_header(ws_rank, cols_rank)
    ws_rank.freeze_panes = "A2"

    row = 2
    lines = ["Central", "Western", "Harbour"]
    periods = ["morning_peak", "evening_peak", "off_peak", "night"]
    for line in lines:
        for period in periods:
            try:
                df = peak_rankings(store, line, period)
            except Exception:
                logger.warning("Rankings fetch failed for %s/%s", line, period, exc_info=True)
                continue
            for r in df.iter_rows(named=True):
                avg = float(r["avg_delay"]) if r["avg_delay"] is not None else 0.0
                ci_l = float(r["ci_lower"]) if r.get("ci_lower") is not None else None
                ci_u = float(r["ci_upper"]) if r.get("ci_upper") is not None else None
                vals = [r["station_name"], line, period.replace("_", " ").title(), avg, ci_l, ci_u]
                for c, v in enumerate(vals, start=1):
                    cell = ws_rank.cell(row=row, column=c, value=v)
                    cell.border = thin_border()
                    if c == 4 and isinstance(v, float):
                        cell.number_format = "0.00"
                        if v >= 5.0:
                            cell.fill = fill(_RED_FILL)
                        elif v >= 3.0:
                            cell.fill = fill(_YELLOW_FILL)
                        else:
                            cell.fill = fill(_GREEN_FILL)
                row += 1

    # Color scale on Avg Delay column (D2:D<last>)
    if row > 2:
        cs_rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws_rank.conditional_formatting.add(f"D2:D{row - 1}", cs_rule)

    auto_width(ws_rank)
    add_meta(ws_rank)

    # ── Sheet 2: Anomalies ──────────────────────────────────────────────────
    ws_anom = wb.create_sheet("Anomalies")
    cols_anom = ["Station", "Severity", "Actual Delay (min)", "Expected (min)", "Upper Bound", "Date"]
    write_header(ws_anom, cols_anom)
    ws_anom.freeze_panes = "A2"

    try:
        import datetime as _dt

        import polars as pl

        from analysis.anomaly import AnomalyBatch

        all_lines = ["Central", "Western", "Harbour"]
        top_stations: list[str] = []
        for line in all_lines:
            try:
                df = store.worst_stations(line, n=5)
                top_stations.extend(df["station_name"].to_list())
            except Exception:
                continue

        history_parts: list[pl.DataFrame] = []
        for station in top_stations:
            try:
                daily = store.daily_avg(station)
                if len(daily) == 0:
                    continue
                daily = daily.with_columns(pl.lit(station).alias("station_name"))
                daily = daily.with_columns(pl.col("date").cast(pl.Utf8).alias("date"))
                history_parts.append(daily)
            except Exception:
                continue

        anom_row = 2
        if history_parts:
            history = pl.concat(history_parts)
            today_parts = []
            for station in top_stations:
                sh = history.filter(pl.col("station_name") == station)
                if len(sh) > 0:
                    today_parts.append(sh.sort("date").tail(1))
            if today_parts:
                today = pl.concat(today_parts)
                batch = AnomalyBatch(history=history)
                results = batch.detect_all(today)
                today_str = str(_dt.date.today())
                for r in results:
                    severity_fill = _RED_FILL if r.severity == "HIGH" else (_YELLOW_FILL if r.severity == "MEDIUM" else _GREEN_FILL)
                    vals = [r.station, r.severity, r.actual_delay, r.expected_delay, r.upper_bound, today_str]
                    for c, v in enumerate(vals, start=1):
                        cell = ws_anom.cell(row=anom_row, column=c, value=v)
                        cell.border = thin_border()
                        if c in (2,):
                            cell.fill = fill(severity_fill)
                        if c in (3, 4, 5) and isinstance(v, float):
                            cell.number_format = "0.00"
                    anom_row += 1
    except Exception:
        logger.warning("Anomaly sheet build failed", exc_info=True)
        ws_anom.cell(row=2, column=1, value="No anomaly data available")

    auto_width(ws_anom)
    add_meta(ws_anom)

    # ── Sheet 3: Line Trends ────────────────────────────────────────────────
    ws_trend = wb.create_sheet("Line Trends")
    cols_trend = ["Date", "Line", "Avg Delay (min)"]
    write_header(ws_trend, cols_trend)
    ws_trend.freeze_panes = "A2"

    trend_row = 2
    for line in lines:
        try:
            df = store.line_trend(line)
            for r in df.iter_rows(named=True):
                vals = [str(r["date"]), line, float(r["avg_delay"]) if r["avg_delay"] is not None else 0.0]
                for c, v in enumerate(vals, start=1):
                    cell = ws_trend.cell(row=trend_row, column=c, value=v)
                    cell.border = thin_border()
                    if c == 3 and isinstance(v, float):
                        cell.number_format = "0.00"
                trend_row += 1
        except Exception:
            logger.warning("Line trend fetch failed for %s", line, exc_info=True)
            continue

    auto_width(ws_trend)
    add_meta(ws_trend)

    # ── Sheet 4: Heatmap (Dadar CR) ─────────────────────────────────────────
    ws_heat = wb.create_sheet("Heatmap (Dadar CR)")
    DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    HOURS = [f"{h:02d}:00" for h in range(24)]

    # Header row: hours
    ws_heat.cell(row=1, column=1, value="Day \\ Hour").font = hdr_font()
    ws_heat.cell(row=1, column=1).fill = hdr_fill()
    for h, label in enumerate(HOURS, start=2):
        cell = ws_heat.cell(row=1, column=h, value=label)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal="center")

    # Day labels + heatmap data
    try:
        df = store.heatmap("Dadar CR")
        matrix: list[list[float | None]] = [[None] * 24 for _ in range(7)]
        for r in df.iter_rows(named=True):
            wd, hr = int(r["weekday"]), int(r["hour"])
            if 0 <= wd < 7 and 0 <= hr < 24:
                v = r["avg_delay"]
                matrix[wd][hr] = float(v) if v is not None else None
    except Exception:
        matrix = [[None] * 24 for _ in range(7)]

    for wd, day in enumerate(DAYS, start=0):
        row_idx = wd + 2
        day_cell = ws_heat.cell(row=row_idx, column=1, value=day)
        day_cell.font = hdr_font()
        day_cell.fill = hdr_fill()
        for hr in range(24):
            val = matrix[wd][hr]
            cell = ws_heat.cell(row=row_idx, column=hr + 2, value=val)
            cell.border = thin_border()
            if val is not None:
                cell.number_format = "0.0"
                if val >= 6.0:
                    cell.fill = fill(_RED_FILL)
                elif val >= 3.5:
                    cell.fill = fill(_YELLOW_FILL)
                else:
                    cell.fill = fill(_GREEN_FILL)

    ws_heat.column_dimensions["A"].width = 6
    for h in range(24):
        ws_heat.column_dimensions[get_column_letter(h + 2)].width = 6
    ws_heat.freeze_panes = "B2"
    add_meta(ws_heat)

    # ── Sheet 5: Summary ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)  # insert at front
    ws_sum.sheet_properties.tabColor = _ACCENT.lstrip("#")

    ws_sum["A1"] = "Mumbai Local Delay Analytics — Executive Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1C3557")
    ws_sum.merge_cells("A1:D1")

    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Value"
    ws_sum["C3"] = "Notes"
    for col in ("A", "B", "C"):
        cell = ws_sum[f"{col}3"]
        cell.font = hdr_font()
        cell.fill = hdr_fill()

    try:
        from analysis.rankings import peak_rankings as _pr
        central_morning = _pr(store, "Central", "morning_peak")
        worst_station = central_morning["station_name"][0] if len(central_morning) > 0 else "N/A"
        worst_delay = float(central_morning["avg_delay"][0]) if len(central_morning) > 0 else 0.0
        metrics = [
            ("Worst Station (Central Morning)", worst_station, "Highest avg delay — Central line AM peak"),
            ("Worst Delay", f"{worst_delay:.1f} min", "Avg delay at worst station"),
            ("Lines Covered", "Central, Western, Harbour", "3 Mumbai rail corridors"),
            ("Data Source", "etrain-delays.in (scraped)", "Real-time delay reports"),
            ("Model", "Prophet time-series + DuckDB", "Anomaly detection + SQL analytics"),
            ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), "Refresh via VBA macro or /api/export/excel"),
        ]
    except Exception:
        metrics = [
            ("Status", "Data unavailable", "Run seed_db.py to populate DuckDB"),
            ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
        ]

    for i, (metric, value, note) in enumerate(metrics, start=4):
        ws_sum.cell(row=i, column=1, value=metric).border = thin_border()
        ws_sum.cell(row=i, column=2, value=value).border = thin_border()
        ws_sum.cell(row=i, column=3, value=note).font = Font(italic=True, color="888888")

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 48

    # ── Finalise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get(
    "/export/excel",
    response_class=StreamingResponse,
    summary="Export all analytics to formatted Excel workbook",
    responses={
        200: {
            "description": "Excel workbook (.xlsx) with 5 sheets: Summary, Rankings, Anomalies, Line Trends, Heatmap",
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
        }
    },
)
def export_excel(store: DelayStore = Depends(get_store)) -> StreamingResponse:  # noqa: B008
    """Generate a formatted .xlsx export of all analytics data.

    Consumed by:
    - React frontend "Export to Excel" button
    - VBA macro in excel/RefreshData.bas (pulls this endpoint on Workbook_Open)
    """
    buf = _build_workbook(store)
    buf.seek(0, 2)
    size = buf.tell()
    buf.seek(0)
    filename = f"mumbai_local_delays_{datetime.date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(size),
        },
    )
